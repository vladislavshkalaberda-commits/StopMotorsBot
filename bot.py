"""
Motor Search Telegram Bot — финальная версия
"""

import logging
import os
from openpyxl import load_workbook
from telegram import Update
from telegram.ext import ApplicationBuilder, CommandHandler, MessageHandler, filters, ContextTypes

# ─── НАСТРОЙКИ ────────────────────────────────────────────────────────────────
BOT_TOKEN = os.environ.get("BOT_TOKEN", "ВСТАВЬ_ТОКЕН_СЮДА")
LISTE_FILE = "LISTE_100_-_31-12-2025.xlsx"
PRICES_FILE = "Моторы_цены_диапазон.xlsx"
MIN_QUERY_LEN = 3

# Коды которые НЕ привязываем к списку 1 (случайные совпадения)
IGNORE_CODES = {"OM", "MB", "CD"}

# Коды у которых поиск только по точному совпадению начала слова (не подстрока)
EXACT_START_CODES = {"Z1", "HR", "YD", "QG", "F14D", "B12D", "B12S", "350A"}

logging.basicConfig(
    format="%(asctime)s - %(name)s - %(levelname)s - %(message)s",
    level=logging.INFO,
)
logger = logging.getLogger(__name__)

# ─── ЗАГРУЗКА ДАННЫХ ──────────────────────────────────────────────────────────

KNOWN_BRANDS = {
    "audi", "bmw", "chevrolet", "citroen", "citroen| peugeot",
    "daewoo", "fiat", "ford", "jeep", "hyundai/kia", "hyundai",
    "kia", "mercedes", "nissan", "opel", "renault", "rover",
    "volkswagen", "vw", "volvo", "peugeot",
}

def load_liste(filepath: str) -> list[dict]:
    """Загружает список 1 — коды моторов с марками."""
    wb = load_workbook(filepath, read_only=True, data_only=True)
    motors = []
    for sheet in wb.worksheets:
        current_brand = "—"
        for row in sheet.iter_rows(values_only=True):
            for cell in row:
                if cell is None:
                    continue
                v = str(cell).strip()
                if not v:
                    continue
                if v.lower() in KNOWN_BRANDS:
                    current_brand = v
                    continue
                if len(v) <= 20 and not any(c.isdigit() for c in v) and v[0].isupper():
                    current_brand = v
                    continue
                motors.append({"brand": current_brand, "code": v})
    wb.close()
    return motors


def load_prices(filepath: str) -> dict[str, dict]:
    """
    Загружает список 2 — коды с ценами.
    Возвращает словарь: { "K9K": {"min": 150, "max": 400, "qty": 420}, ... }
    """
    wb = load_workbook(filepath, read_only=True, data_only=True)
    prices = {}
    for sheet in wb.worksheets:
        header_passed = False
        for row in sheet.iter_rows(values_only=True):
            if not header_passed:
                if row[1] and "Код" in str(row[1]):
                    header_passed = True
                continue
            if row[1] is None:
                continue
            code = str(row[1]).strip()
            if not code or code == "ИТОГО":
                continue
            if code in IGNORE_CODES:
                continue
            try:
                qty = int(row[2]) if row[2] else 0
                mn  = int(row[3]) if row[3] else None
                mx  = int(row[4]) if row[4] else None
            except (ValueError, TypeError):
                continue
            prices[code] = {"min": mn, "max": mx, "qty": qty}
    wb.close()
    return prices


# Глобальные данные
LISTE: list[dict] = []
PRICES: dict[str, dict] = {}


def reload_data():
    global LISTE, PRICES
    base = os.path.dirname(os.path.abspath(__file__))
    liste_path  = os.path.join(base, LISTE_FILE)
    prices_path = os.path.join(base, PRICES_FILE)

    if os.path.exists(liste_path):
        LISTE = load_liste(liste_path)
        logger.info(f"Список 1 загружен: {len(LISTE)} моторов")
    else:
        logger.error(f"Файл не найден: {liste_path}")

    if os.path.exists(prices_path):
        PRICES = load_prices(prices_path)
        logger.info(f"Список 2 загружен: {len(PRICES)} кодов с ценами")
    else:
        logger.error(f"Файл не найден: {prices_path}")


# ─── ЛОГИКА ПОИСКА ────────────────────────────────────────────────────────────

def matches_code(query: str, code: str) -> bool:
    """
    Проверяет, совпадает ли запрос с кодом из списка 1.

    Правила:
    - Для кодов из EXACT_START_CODES: совпадение только если слово в коде
      начинается точно с query (например, HR → HR16, но не 4HR).
    - Для остальных: query является подстрокой кода.
    """
    q = query.upper()
    c = code.upper()

    if q in EXACT_START_CODES:
        # Ищем слово в коде, которое начинается с q
        for word in c.replace(",", " ").split():
            if word.startswith(q):
                return True
        return False
    else:
        return q in c


def matches_price_code(query: str, price_code: str) -> bool:
    """
    Проверяет совпадение запроса с кодом из списка 2 (цены).

    Логика:
    - Запрос совпадает если price_code начинается с query (K9K → K9K 724)
      ИЛИ query начинается с price_code (M57D30 → M57D).
    - Для Z1: только точное совпадение.
    """
    q = query.upper().strip()
    p = price_code.upper().strip()

    if q == "Z1":
        return p == "Z1"

    return p.startswith(q) or q.startswith(p)


def search(query: str) -> dict:
    """
    Возвращает:
      {
        "liste_hits":  [ {"brand": ..., "code": ...}, ... ],  # из списка 1
        "price_info":  {"min": ..., "max": ..., "qty": ...} | None,
        "price_code":  "K9K" | None,   # какой код из списка 2 сработал
      }
    """
    q = query.strip()

    # --- Список 1 ---
    liste_hits = [m for m in LISTE if matches_code(q, m["code"])]

    # --- Список 2 ---
    price_info = None
    price_code_matched = None

    # Сначала пробуем точное совпадение с кодом из списка 2
    for pcode, pdata in PRICES.items():
        if matches_price_code(q, pcode):
            # Берём первое совпадение (или наиболее длинное)
            if price_code_matched is None or len(pcode) > len(price_code_matched):
                price_code_matched = pcode
                price_info = pdata

    return {
        "liste_hits": liste_hits,
        "price_info": price_info,
        "price_code": price_code_matched,
    }


# ─── ФОРМАТИРОВАНИЕ ОТВЕТА ────────────────────────────────────────────────────

def format_response(query: str, result: dict) -> str:
    hits   = result["liste_hits"]
    pinfo  = result["price_info"]
    pcode  = result["price_code"]

    lines = []

    if hits:
        # Группируем по марке
        grouped: dict[str, list[str]] = {}
        for m in hits:
            grouped.setdefault(m["brand"], []).append(m["code"])

        lines.append(f"✅ <b>{query.upper()}</b> — найден в списке ({len(hits)} вар.):\n")
        for brand, codes in grouped.items():
            lines.append(f"🚗 <b>{brand}</b>")
            for code in codes:
                lines.append(f"   • {code}")
        lines.append("")
    else:
        lines.append(f"❌ <b>{query.upper()}</b> — нет в текущем списке.\n")

    if pinfo:
        mn  = pinfo["min"]
        mx  = pinfo["max"]
        qty = pinfo["qty"]
        tag = f" (как <code>{pcode}</code>)" if pcode and pcode.upper() != query.upper() else ""
        if hits:
            lines.append(f"💰 История покупок{tag}: <b>{mn}€ — {mx}€</b>  |  куплено: {qty} шт.")
        else:
            lines.append(f"⚠️ В текущем списке нет, но ранее покупался{tag}:")
            lines.append(f"💰 Диапазон цен: <b>{mn}€ — {mx}€</b>  |  куплено: {qty} шт.")
    else:
        if not hits:
            lines.append("📭 История покупок тоже пуста.")

    return "\n".join(lines).strip()


# ─── ХЕНДЛЕРЫ ─────────────────────────────────────────────────────────────────

async def cmd_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = (
        "👋 Привет! Введи код или маркировку мотора (минимум 3 символа).\n\n"
        "Примеры: <code>K4M</code>, <code>K9K</code>, <code>M57D</code>, <code>N52</code>\n\n"
        "/help — помощь\n"
        "/reload — перезагрузить списки"
    )
    await update.message.reply_text(text, parse_mode="HTML")


async def cmd_help(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = (
        "🔍 <b>Как пользоваться:</b>\n"
        "Напиши код мотора или его часть — минимум 3 символа.\n"
        "Регистр не важен.\n\n"
        "<b>Что покажет бот:</b>\n"
        "• Все совпадения из текущего списка\n"
        "• Диапазон цен из истории покупок\n"
        "• Если мотора нет в списке — сообщит и всё равно покажет цены\n\n"
        f"📋 Моторов в списке: <b>{len(LISTE)}</b>\n"
        f"💰 Кодов с ценами: <b>{len(PRICES)}</b>"
    )
    await update.message.reply_text(text, parse_mode="HTML")


async def cmd_reload(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text("🔄 Перезагружаю списки...")
    reload_data()
    await update.message.reply_text(
        f"✅ Готово!\n📋 Моторов в списке: {len(LISTE)}\n💰 Кодов с ценами: {len(PRICES)}"
    )


async def handle_message(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.message.text.strip()

    if len(query) < MIN_QUERY_LEN:
        await update.message.reply_text(
            f"⚠️ Введи минимум {MIN_QUERY_LEN} символа."
        )
        return

    result = search(query)
    text = format_response(query, result)

    if len(text) > 4000:
        text = text[:3990] + "\n\n...список обрезан, уточни запрос."

    await update.message.reply_text(text, parse_mode="HTML")


# ─── ЗАПУСК ───────────────────────────────────────────────────────────────────

def main():
    reload_data()
    app = ApplicationBuilder().token(BOT_TOKEN).build()
    app.add_handler(CommandHandler("start", cmd_start))
    app.add_handler(CommandHandler("help", cmd_help))
    app.add_handler(CommandHandler("reload", cmd_reload))
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle_message))
    logger.info("Бот запущен...")
    app.run_polling()


if __name__ == "__main__":
    main()
